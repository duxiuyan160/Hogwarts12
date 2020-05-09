import openpyxl
from openpyxl import Workbook, load_workbook


def test_writeexcel():
    wb = Workbook()
    dest_name = "demo.xlsx"
    ws1 = wb.active
    ws1.title = "aa"
    ws1.cell(row=4,column=5,value="duxiuyan")
    ws2 = wb.create_sheet(title="bb")
    wb.save(filename=dest_name)
    wb.close()

def test_readexcel():
    wb = load_workbook(filename="demo.xlsx")
    sheet_ranges = wb["aa"]
    print(sheet_ranges["D12"].value)


